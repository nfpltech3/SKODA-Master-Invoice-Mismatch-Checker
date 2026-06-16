import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import re
import logging
import traceback
from decimal import Decimal, InvalidOperation
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import sys
from pathlib import Path

# ═══════════════════════════════════════════════════════════════
# LOGGING SETUP
# ═══════════════════════════════════════════════════════════════
from logging.handlers import RotatingFileHandler

VERBOSE_LOG = False

_LOG_FILE = Path(__file__).parent / "mismatch_checker.log" if not getattr(sys, 'frozen', False) else Path(sys.executable).parent / "mismatch_checker.log"

_formatter = logging.Formatter('%(asctime)s | %(levelname)-7s | %(message)s')
_fh = RotatingFileHandler(_LOG_FILE, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
_fh.setFormatter(_formatter)

logging.basicConfig(
    level=logging.DEBUG if VERBOSE_LOG else logging.INFO,
    handlers=[_fh]
)
logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
# NORMALIZATION FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def normalize_part_number(val) -> str:
    if pd.isna(val) or val == "": return ""
    v = str(val).strip()
    if v.endswith('.0'): v = v[:-2]
    return v.lower()

def first_two_hs(val) -> str:
    if pd.isna(val) or val == "":
        return ""
    v = str(val).strip().upper()
    if v.endswith(".0"):
        v = v[:-2]
    v = re.sub(r"[^\d]", "", v)
    return v[:2]

def normalize_number(val):
    if pd.isna(val) or val == "": return None
    v = re.sub(r'[^\d.-]', '', str(val).strip())
    if not v or v in ('.', '-'): return None
    try:
        return Decimal(str(float(v))).quantize(Decimal('0.0001'))
    except (InvalidOperation, ValueError):
        return None

def normalize_invoice_number(val) -> str:
    if pd.isna(val) or val == "": return ""
    v = str(val).strip()
    if v.endswith('.0'): v = v[:-2]
    return v.lower()

def normalize_date(val) -> str:
    if pd.isna(val) or val == "": return ""
    if isinstance(val, datetime): return val.strftime('%d-%m-%Y')
    v = str(val).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%y', '%d-%b-%Y']:
        try: return datetime.strptime(v, fmt).strftime('%d-%m-%Y')
        except ValueError: pass
    return v

def normalize_country(val) -> str:
    if pd.isna(val) or val == "": return ""
    v = str(val).strip().upper()
    m = {
        # Europe
        'DE': 'GERMANY', 'HU': 'HUNGARY', 'CZ': 'CZECH REPUBLIC', 'SK': 'SLOVAKIA',
        'XS': 'GERMANY', 'PL': 'POLAND', 'FR': 'FRANCE', 'IT': 'ITALY', 'ES': 'SPAIN',
        'PT': 'PORTUGAL', 'NL': 'NETHERLANDS', 'BE': 'BELGIUM', 'AT': 'AUSTRIA',
        'CH': 'SWITZERLAND', 'SE': 'SWEDEN', 'NO': 'NORWAY', 'DK': 'DENMARK',
        'FI': 'FINLAND', 'IE': 'IRELAND', 'GB': 'UNITED KINGDOM', 'UK': 'UNITED KINGDOM',
        'RO': 'ROMANIA', 'BG': 'BULGARIA', 'HR': 'CROATIA', 'SI': 'SLOVENIA',
        'RS': 'SERBIA', 'BA': 'BOSNIA AND HERZEGOVINA', 'ME': 'MONTENEGRO',
        'MK': 'NORTH MACEDONIA', 'AL': 'ALBANIA', 'LT': 'LITHUANIA',
        'LV': 'LATVIA', 'EE': 'ESTONIA', 'LU': 'LUXEMBOURG', 'MT': 'MALTA',
        'CY': 'CYPRUS', 'GR': 'GREECE', 'IS': 'ICELAND', 'UA': 'UKRAINE',
        'BY': 'BELARUS', 'MD': 'MOLDOVA', 'RU': 'RUSSIA',
        # Asia
        'CN': 'CHINA', 'IN': 'INDIA', 'JP': 'JAPAN', 'KR': 'SOUTH KOREA',
        'TW': 'TAIWAN', 'TAIWAN, PROVINCE OF CHINA': 'TAIWAN', 'TH': 'THAILAND', 'VN': 'VIETNAM', 'MY': 'MALAYSIA',
        'SG': 'SINGAPORE', 'ID': 'INDONESIA', 'PH': 'PHILIPPINES', 'BD': 'BANGLADESH',
        'PK': 'PAKISTAN', 'LK': 'SRI LANKA', 'NP': 'NEPAL', 'MM': 'MYANMAR',
        'KH': 'CAMBODIA', 'LA': 'LAOS', 'MN': 'MONGOLIA', 'KZ': 'KAZAKHSTAN',
        'UZ': 'UZBEKISTAN', 'AF': 'AFGHANISTAN',
        # Middle East
        'AE': 'UNITED ARAB EMIRATES', 'SA': 'SAUDI ARABIA', 'QA': 'QATAR',
        'KW': 'KUWAIT', 'BH': 'BAHRAIN', 'OM': 'OMAN', 'IR': 'IRAN',
        'IQ': 'IRAQ', 'IL': 'ISRAEL', 'JO': 'JORDAN', 'LB': 'LEBANON',
        'TR': 'TURKEY', 'GE': 'GEORGIA', 'AM': 'ARMENIA', 'AZ': 'AZERBAIJAN',
        # Americas
        'US': 'UNITED STATES', 'CA': 'CANADA', 'MX': 'MEXICO',
        'BR': 'BRAZIL', 'AR': 'ARGENTINA', 'CL': 'CHILE', 'CO': 'COLOMBIA',
        'PE': 'PERU', 'VE': 'VENEZUELA', 'EC': 'ECUADOR', 'UY': 'URUGUAY',
        'PY': 'PARAGUAY', 'BO': 'BOLIVIA', 'CR': 'COSTA RICA', 'PA': 'PANAMA',
        'DO': 'DOMINICAN REPUBLIC', 'GT': 'GUATEMALA', 'CU': 'CUBA',
        'HN': 'HONDURAS', 'SV': 'EL SALVADOR', 'NI': 'NICARAGUA',
        # Africa
        'ZA': 'SOUTH AFRICA', 'EG': 'EGYPT', 'NG': 'NIGERIA', 'KE': 'KENYA',
        'MA': 'MOROCCO', 'TN': 'TUNISIA', 'GH': 'GHANA', 'ET': 'ETHIOPIA',
        'TZ': 'TANZANIA', 'DZ': 'ALGERIA', 'LY': 'LIBYA',
        # Oceania
        'AU': 'AUSTRALIA', 'NZ': 'NEW ZEALAND',
    }
    # Also build reverse mapping (full name -> full name) so both sides normalize consistently
    reverse = {name: name for name in m.values()}
    combined = {**m, **reverse}
    return combined.get(v, v)

def normalize_description(val, model="", generic_desc="") -> str:
    """Normalize description by removing model number and generic description text."""
    if pd.isna(val) or val == "": return ""
    v = str(val).upper().strip()
    # Remove the model number from the description
    if model and not pd.isna(model):
        v = v.replace(str(model).upper().strip(), '')
    # Remove the generic description (from separate column in Item Report)
    if generic_desc and not pd.isna(generic_desc):
        gd = str(generic_desc).upper().strip()
        v = v.replace(gd, '')
    # Remove text inside brackets (any remaining generic descriptions)
    v = re.sub(r'\(.*?\)', '', v)
    # Ignore punctuation and extra spaces
    v = re.sub(r'[^\w\s]', '', v)
    return re.sub(r'\s+', ' ', v).strip()

def normalize_unit(ext_price_per_pc: str) -> str:
    """Extract unit from Price per PC (e.g. '100 /PC' -> 'PCS', '100 /G' -> 'GMS', '100 /L' -> 'LTR')"""
    if pd.isna(ext_price_per_pc) or ext_price_per_pc == "": return ""
    s = str(ext_price_per_pc).upper().strip()
    if '/' not in s:
        return ""
    unit_part = s.split('/')[-1].strip().replace(" ", "")
    if unit_part == 'PC': return 'PCS'
    if unit_part == 'G': return 'GMS'
    if unit_part == 'L': return 'LTR'
    if unit_part == 'KG': return 'KGS'
    return unit_part

# ═══════════════════════════════════════════════════════════════
# COMPARISON LOGIC
# ═══════════════════════════════════════════════════════════════

def validate_columns(df, required_cols: list[str], file_name: str, sheet_name: str) -> None:
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns in {file_name} / {sheet_name}:\n" + "\n".join(f"- {m}" for m in missing))

def load_dataframe(path: str, sheet_name=0, dtype=None) -> pd.DataFrame:
    """Helper to load either Excel or CSV transparently."""
    p = str(path).lower()
    if p.endswith('.csv'):
        try:
            return pd.read_csv(path, dtype=dtype, encoding='utf-8-sig')
        except UnicodeDecodeError:
            return pd.read_csv(path, dtype=dtype, encoding='latin1')
    return pd.read_excel(path, sheet_name=sheet_name, dtype=dtype)

def perform_comparison(master_path: str, item_path: str, ext_path: str, status_callback=None) -> dict:
    """Run comparison and return dict with mismatch DF and summary counts."""
    if status_callback: status_callback("Loading files...")
    logger.info("Starting comparison...")
    logger.info("Master: %s | Item: %s | Invoice: %s", master_path, item_path, ext_path)
    try: 
        df_master = load_dataframe(master_path, sheet_name="Master Sheet", dtype=str)
        logger.info("Master Sheet loaded | Shape: %s | Cols: %s | First 3 PartNo: %s", df_master.shape, list(df_master.columns), df_master.get('PartNo', pd.Series(dtype=str)).head(3).tolist())
    except Exception as e: logger.error("Master Sheet load failed: %s", e); raise
    try: 
        df_item = load_dataframe(item_path, sheet_name=0)
        df_item['excel_row_item'] = df_item.index + 2
        logger.info("Item Report loaded | Shape: %s | Cols: %s | First 3 Model: %s", df_item.shape, list(df_item.columns), df_item.get('Model', pd.Series(dtype=str)).head(3).tolist())
    except Exception as e: logger.error("Item Report load failed: %s", e); raise
    try: 
        df_ext = load_dataframe(ext_path, sheet_name=0)
        df_ext['excel_row_ext'] = df_ext.index + 2
        ext_key_col = 'Mat. NO.' if 'Mat. NO.' in df_ext.columns else 'Part Number' if 'Part Number' in df_ext.columns else df_ext.columns[0]
        logger.info("Extracted Invoice loaded | Shape: %s | Cols: %s | First 3 %s: %s", df_ext.shape, list(df_ext.columns), ext_key_col, df_ext.get(ext_key_col, pd.Series(dtype=str)).head(3).tolist())
    except Exception as e: logger.error("Extracted Invoice load failed: %s", e); raise

    # ── Column aliasing for alternate invoice formats ──
    # Canonical names used internally vs alternate names found in different invoice layouts.
    ext_aliases = {
        'Part Number': 'Mat. NO.',
        'Part No.': 'Mat. NO.',
        'Article No': 'Mat. NO.',
        'Your Reference': 'Mat. NO.',
        'COO': 'Country of Origin',
        'Country Code': 'Country of Origin',
        'QUANTITY': 'Quantity',
        'Qty': 'Quantity',
        'VALUE OF GOODS': 'Total Price',
        'Amount': 'Total Price',
        'HS-CODE': 'HS Code',
        'Invoice No': 'Invoice Number',
    }
    unmatched_aliases = []
    for alt, canonical in ext_aliases.items():
        if alt in df_ext.columns and canonical not in df_ext.columns:
            df_ext.rename(columns={alt: canonical}, inplace=True)
            logger.info("Aliased Extracted Invoice column '%s' -> '%s'", alt, canonical)
        elif alt not in df_ext.columns:
            # Alt name not present in file at all — truly unmatched
            unmatched_aliases.append(alt)
        # If canonical already exists, alias is irrelevant — no need to log
    logger.info("Column aliases attempted but did not match: %s", unmatched_aliases)

    master_req = ['PartNo', 'CTH1', 'Basic Duty Rate', 'IGST Notn Sr No']
    item_req = ['Job No','Job Date','Invoice No','Invoice Date','Model','Product Desc','CTH','Quantity','Amount','Country of Origin','Basic Duty Rate','IGST Notification SrNo']
    ext_req = ['Invoice Number','Invoice Date','Country of Origin','Mat. NO.','Description','Quantity','Total Price']

    validate_columns(df_master, master_req, "Master Sheet", "Master Sheet")
    validate_columns(df_item, item_req, "Item Report", "First Sheet")
    validate_columns(df_ext, ext_req, "Extracted Invoice", "First Sheet")
    logger.info("Column validation passed. Master=%d rows, Item=%d rows, Invoice=%d rows", len(df_master), len(df_item), len(df_ext))
    if status_callback: status_callback(f"Normalizing data ({len(df_item) + len(df_ext)} rows)...")

    # Strip unnamed columns from both dataframes
    df_item = df_item[[c for c in df_item.columns if 'Unnamed' not in str(c)]]
    df_ext = df_ext[[c for c in df_ext.columns if 'Unnamed' not in str(c)]]

    # Filter out summary/blank rows from Extracted Invoice
    # Rows where Mat. NO. is blank are summary lines (e.g. "Total Unique Packages: 113")
    before_filter_df = df_ext.copy()
    df_ext = df_ext.dropna(subset=['Mat. NO.'])
    df_ext = df_ext[df_ext['Mat. NO.'].astype(str).str.strip() != '']
    dropped = len(before_filter_df) - len(df_ext)
    if dropped > 0:
        dropped_df = before_filter_df[~before_filter_df.index.isin(df_ext.index)]
        first_3_dropped = dropped_df['Mat. NO.'].head(3).tolist()
        logger.info("Filtered %d summary/blank rows from Extracted Invoice. First 3 dropped Mat. NO.: %s", dropped, first_3_dropped)

    # Normalize match keys
    df_master['match_key'] = df_master['PartNo'].apply(normalize_part_number)
    df_item['match_key'] = df_item['Model'].apply(normalize_part_number)
    df_ext['match_key'] = df_ext['Mat. NO.'].apply(normalize_part_number)

    mk_master = set(df_master['match_key'].unique())
    mk_item = set(df_item['match_key'].unique())
    mk_ext = set(df_ext['match_key'].unique())
    logger.info("Match Keys Unique Counts - Master: %d, Item: %d, Extracted: %d", len(mk_master), len(mk_item), len(mk_ext))
    logger.info("Match Keys ONLY in Item Report (max 20): %s", list(mk_item - mk_ext)[:20])
    logger.info("Match Keys ONLY in Extracted Invoice (max 20): %s", list(mk_ext - mk_item)[:20])
    logger.info("Match Keys missing from BOTH Invoice and Master (max 20): %s", list(mk_item - mk_ext - mk_master)[:20])

    # Master dedup & conflict detection
    master_conflicts = set()
    df_master_unique = df_master.drop_duplicates(subset=['match_key'], keep='first').copy()
    for key, grp in df_master.groupby('match_key'):
        if len(grp) > 1 and (grp['CTH1'].nunique() > 1 or grp['Basic Duty Rate'].nunique() > 1 or grp['IGST Notn Sr No'].nunique() > 1):
            master_conflicts.add(key)
    df_master_unique['Master_Conflict'] = df_master_unique['match_key'].isin(master_conflicts)

    # Occurrence index for duplicate handling
    df_item['occurrence'] = df_item.groupby('match_key').cumcount()
    df_ext['occurrence'] = df_ext.groupby('match_key').cumcount()
    item_counts = df_item['match_key'].value_counts()
    ext_counts = df_ext['match_key'].value_counts()

    # Merge
    if status_callback: status_callback("Merging and identifying mismatches...")
    logger.info("Merging Item Report (%d) with Extracted Invoice (%d)...", len(df_item), len(df_ext))
    merged = pd.merge(df_item, df_ext, on=['match_key','occurrence'], how='outer', suffixes=('_item','_ext'), indicator=True)
    logger.debug("Merged columns: %s", list(merged.columns))
    
    merge_counts = merged['_merge'].value_counts().to_dict()
    logger.info("Merge distribution (_merge counts): %s", merge_counts)
    
    item_dupes = df_item[df_item['occurrence'] > 0]['match_key'].unique()
    ext_dupes = df_ext[df_ext['occurrence'] > 0]['match_key'].unique()
    if len(item_dupes) > 0:
        logger.info("Duplicate match keys in Item Report (max 20): %s", list(item_dupes)[:20])
    if len(ext_dupes) > 0:
        logger.info("Duplicate match keys in Extracted Invoice (max 20): %s", list(ext_dupes)[:20])
        
    merged_full = pd.merge(merged, df_master_unique, on='match_key', how='left')

    mismatches: list[dict] = []
    cth_alerts: list[dict] = []

    def add(rd, field, m_val, e_val, i_val, remark, status="MISMATCH", sub_row=False, target_list=None, inv_cth="", item_cth="", inv_desc="", item_desc=""):
        if target_list is None: target_list = mismatches
        def clean_inv(v):
            if pd.isna(v) or v == "": return ""
            s = str(v).strip()
            if s.endswith('.0'): s = s[:-2]
            return s

        # Append row context to remarks
        r_item = rd.get('excel_row_item')
        r_ext = rd.get('excel_row_ext')
        row_info = []
        if pd.notna(r_item) and r_item != "": row_info.append(f"Item Row: {int(r_item)}")
        if pd.notna(r_ext) and r_ext != "": row_info.append(f"Inv Row: {int(r_ext)}")
        if row_info:
            remark = f"{remark} [{', '.join(row_info)}]"

        rec = {
            'Job No': rd.get('Job No','') if not sub_row else '', 
            'Job Date': normalize_date(rd.get('Job Date','')) if not sub_row else '',
            'Invoice No': clean_inv(rd.get('Invoice No','')) if not sub_row else '', 
            'Invoice Date': normalize_date(rd.get('Invoice Date_item','')) if not sub_row else '',
            'Model': rd.get('Model', rd.get('Mat. NO.','')) if not sub_row else '',
            'Status': status,
            'Mismatch Field': field, 'Master Sheet Value': m_val,
            'Extracted Invoice Value': e_val, 'Item Report Value': i_val,
            'Invoice CTH': inv_cth, 'Item CTH': item_cth,
            'Invoice Desc': inv_desc, 'Item Desc': item_desc,
            'Remarks': remark
        }
        if not sub_row:
            if pd.isna(rec['Invoice No']) or rec['Invoice No'] == '': rec['Invoice No'] = clean_inv(rd.get('Invoice Number',''))
            if pd.isna(rec['Invoice Date']) or rec['Invoice Date'] == '': rec['Invoice Date'] = normalize_date(rd.get('Invoice Date_ext',''))
        target_list.append(rec)

    for _, row in merged_full.iterrows():
        ms = row['_merge']
        mk = row['match_key']

        # Missing in both Invoice and Master — single combined row
        if ms == 'left_only' and pd.isna(row.get('PartNo')) and mk != "":
            add(
                row,
                'Model',
                '',
                '',
                row.get('Model', ''),
                'Model found in Item Report but missing in both Extracted Invoice and Master Sheet'
            )
            continue

        # Missing / Duplicate count
        if ms == 'left_only':
            if mk in ext_counts.index:
                i_count = item_counts.get(mk, 0)
                e_count = ext_counts.get(mk, 0)
                rmk = f'Item Report contains {i_count} row(s) for this Model, but Extracted Invoice contains only {e_count} row(s).'
            else:
                rmk = 'Model found in Item Report but missing in Extracted Invoice'
            add(row, 'Duplicate Occurrence', '', '', row['Model'], rmk)
        elif ms == 'right_only':
            if mk in item_counts.index:
                i_count = item_counts.get(mk, 0)
                e_count = ext_counts.get(mk, 0)
                rmk = f'Extracted Invoice contains {e_count} row(s) for this Model, but Item Report contains only {i_count} row(s).'
            else:
                rmk = 'Model found in Extracted Invoice but missing in Item Report'
            add(row, 'Duplicate Occurrence', '', row['Mat. NO.'], '', rmk)

        # Master presence
        if pd.isna(row.get('PartNo')) and mk != "":
            ext_cth = str(row.get('HS Code', '')).strip()
            item_cth = str(row.get('CTH', '')).strip()
            if ext_cth.endswith('.0'): ext_cth = ext_cth[:-2]
            if item_cth.endswith('.0'): item_cth = item_cth[:-2]
            
            ext_desc = str(row.get('Description_x', '')).strip()
            item_desc = str(row.get('Product Desc', '')).strip()
            
            add(
                row,
                'Master Reference',
                '',
                row.get('Mat. NO.', ''),
                row.get('Model', ''),
                'Model missing in Master Sheet',
                inv_cth=ext_cth,
                item_cth=item_cth,
                inv_desc=ext_desc,
                item_desc=item_desc
            )

        elif row.get('Master_Conflict') == True:
            add(
                row,
                'PartNo',
                'Multiple conflicting values',
                '',
                '',
                'Duplicate Part Number in Master Sheet with conflicting reference values'
            )

        # Master vs Item Report fields
        if ms in ('both','left_only') and not pd.isna(row.get('PartNo')):
            mc = str(row.get('CTH1','')).strip(); ic = str(row.get('CTH','')).strip()
            if mc.endswith('.0'): mc = mc[:-2]
            if ic.endswith('.0'): ic = ic[:-2]
            if mc != ic and mc != 'nan' and ic != 'nan':
                logger.debug("Mismatch [CTH] key: %s | Raw Master: %r, Item: %r | Norm Master: %r, Item: %r", mk, row.get('CTH1',''), row.get('CTH',''), mc, ic)
                add(row, 'CTH', row.get('CTH1',''), '', row.get('CTH',''), 'CTH mismatch between Master Sheet and Item Report')

            # Basic Duty Rate — after outer merge Item's col stays as 'Basic Duty Rate',
            # after left-joining Master which also has 'Basic Duty Rate', pandas adds _x/_y
            m_bdr = row.get('Basic Duty Rate_y', ''); i_bdr = row.get('Basic Duty Rate_x', '')
            norm_m_bdr = normalize_number(m_bdr); norm_i_bdr = normalize_number(i_bdr)
            if norm_m_bdr != norm_i_bdr and not (norm_m_bdr is None and norm_i_bdr is None):
                logger.debug("Mismatch [Basic Duty Rate] key: %s | Raw Master: %r, Item: %r | Norm Master: %r, Item: %r", mk, m_bdr, i_bdr, norm_m_bdr, norm_i_bdr)
                add(row, 'Basic Duty Rate', m_bdr, '', i_bdr, 'Basic Duty Rate mismatch between Master Sheet & Item Report')

            mi = str(row.get('IGST Notn Sr No','')).strip(); ii = str(row.get('IGST Notification SrNo','')).strip()
            if mi.endswith('.0'): mi = mi[:-2]
            if ii.endswith('.0'): ii = ii[:-2]
            if mi != ii and mi != 'nan' and ii != 'nan':
                logger.debug("Mismatch [IGST Notification SrNo] key: %s | Raw Master: %r, Item: %r | Norm Master: %r, Item: %r", mk, row.get('IGST Notn Sr No',''), row.get('IGST Notification SrNo',''), mi, ii)
                add(row, 'IGST Notification SrNo', row.get('IGST Notn Sr No',''), '', row.get('IGST Notification SrNo',''), 'IGST Notification SrNo mismatch between Master Sheet & Item Report')

            # Part Suffix (Master) vs Generic Description (Item Report)
            m_suffix = str(row.get('Part Suffix', '')).strip()
            i_generic = str(row.get('Generic Description', '')).strip()
            if m_suffix and i_generic and m_suffix != 'nan' and i_generic != 'nan':
                if m_suffix.upper() != i_generic.upper():
                    logger.debug("Mismatch [Generic Description] key: %s | Raw Master: %r, Item: %r | Norm Master: %r, Item: %r", mk, row.get('Part Suffix', ''), row.get('Generic Description', ''), m_suffix.upper(), i_generic.upper())
                    add(row, 'Generic Description', m_suffix, '', i_generic, 'Gen. Desc. does not match in Item Report & Master Sheet.')

        # Invoice-level comparisons (only when both sides present)
        if ms == 'both':
            model_val = row.get('Model', '')
            generic_desc_val = row.get('Generic Description', '')
            # Master Sheet description for this part (Description_y after joining master)
            master_desc_raw = row.get('Description_y', '')
            # Extracted Invoice description (Description_x after outer merge with _ext suffix not applied since Item Report has no Description col)
            ext_desc_raw = row.get('Description_x', '')
            # Item Report description
            item_desc_raw = row.get('Product Desc', '')

            e_desc_norm = normalize_description(ext_desc_raw, model=model_val)
            i_desc_norm = normalize_description(item_desc_raw, model=model_val, generic_desc=generic_desc_val)
            if e_desc_norm != i_desc_norm:
                logger.debug("Mismatch [Description] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, ext_desc_raw, item_desc_raw, e_desc_norm, i_desc_norm)
                add(row, 'Description', master_desc_raw, ext_desc_raw, item_desc_raw, 'Description mismatch')
            
            raw_ext_inv = row.get('Invoice Number','')
            raw_item_inv = row.get('Invoice No','')
            norm_ext_inv = normalize_invoice_number(raw_ext_inv)
            norm_item_inv = normalize_invoice_number(raw_item_inv)
            if norm_ext_inv != norm_item_inv:
                logger.debug("Mismatch [Invoice Number] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, raw_ext_inv, raw_item_inv, norm_ext_inv, norm_item_inv)
                add(row, 'Invoice Number', '', raw_ext_inv, raw_item_inv, 'Invoice Number mismatch')
            
            raw_ext_date = row.get('Invoice Date_ext','')
            raw_item_date = row.get('Invoice Date_item','')
            norm_ext_date = normalize_date(raw_ext_date)
            norm_item_date = normalize_date(raw_item_date)
            if norm_ext_date != norm_item_date:
                logger.debug("Mismatch [Invoice Date] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, raw_ext_date, raw_item_date, norm_ext_date, norm_item_date)
                add(row, 'Invoice Date', '', raw_ext_date, raw_item_date, 'Invoice Date mismatch')
            
            raw_ext_coo = row.get('Country of Origin_ext','')
            raw_item_coo = row.get('Country of Origin_item','')
            norm_ext_coo = normalize_country(raw_ext_coo)
            norm_item_coo = normalize_country(raw_item_coo)
            if norm_ext_coo != norm_item_coo:
                logger.debug("Mismatch [Country of Origin] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, raw_ext_coo, raw_item_coo, norm_ext_coo, norm_item_coo)
                add(row, 'Country of Origin', '', raw_ext_coo, raw_item_coo, 'Country of Origin mismatch')
            
            raw_ext_qty = row.get('Quantity_ext','')
            raw_item_qty = row.get('Quantity_item','')
            norm_ext_qty = normalize_number(raw_ext_qty)
            norm_item_qty = normalize_number(raw_item_qty)
            if norm_ext_qty != norm_item_qty:
                logger.debug("Mismatch [Quantity] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, raw_ext_qty, raw_item_qty, norm_ext_qty, norm_item_qty)
                add(row, 'Quantity', '', raw_ext_qty, raw_item_qty, 'Quantity mismatch')
            
            raw_ext_amt = row.get('Total Price','')
            raw_item_amt = row.get('Amount','')
            norm_ext_amt = normalize_number(raw_ext_amt)
            norm_item_amt = normalize_number(raw_item_amt)
            if norm_ext_amt != norm_item_amt:
                logger.debug("Mismatch [Amount] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, raw_ext_amt, raw_item_amt, norm_ext_amt, norm_item_amt)
                add(row, 'Amount', '', raw_ext_amt, raw_item_amt, 'Amount & Total Price mismatch')
                
            # Unit vs Price per PC check
            raw_ext_price_pc = row.get('Price per PC_ext', row.get('Price per PC', ''))
            raw_item_unit = row.get('Unit_item', row.get('Unit', ''))
            norm_ext_unit = normalize_unit(raw_ext_price_pc)
            norm_item_unit = str(raw_item_unit).upper().strip() if pd.notna(raw_item_unit) and str(raw_item_unit).strip() else ""
            
            if norm_ext_unit and norm_item_unit and norm_ext_unit != norm_item_unit:
                logger.debug("Mismatch [Unit] key: %s | Raw Ext: %r, Item: %r | Norm Ext: %r, Item: %r", mk, raw_ext_price_pc, raw_item_unit, norm_ext_unit, norm_item_unit)
                add(row, 'Unit', '', raw_ext_price_pc, raw_item_unit, 'Unit does not match in Item Report & Invoice')

        # Check for CTH Alerts
        if ms in ('both', 'left_only'):
            item_cth = str(row.get('CTH', '')).strip()
            if item_cth.endswith('.0'):
                item_cth = item_cth[:-2]
            
            target_prefixes = ('74', '76', '40111010', '85261000', '9401')
            if any(item_cth.startswith(prefix) for prefix in target_prefixes):
                add(
                    row,
                    field='CTH Alert',
                    m_val='',
                    e_val='',
                    i_val=item_cth,
                    remark='CTH Code Detected',
                    status='CTH_ALERT',
                    sub_row=False,
                    target_list=cth_alerts
                )

    mismatches.extend(cth_alerts)
    
    unique_jobs = []
    if 'Job No' in df_item.columns:
        unique_jobs = df_item['Job No'].dropna().astype(str).unique().tolist()
        unique_jobs = [j for j in unique_jobs if j.strip() != '']

    actual_mismatches = [m for m in mismatches if m.get('Status') not in ('CTH_ALERT', 'INFO')]
    if not actual_mismatches:
        for job in unique_jobs:
            mismatches.append({
                'Job No': job,
                'Job Date': '',
                'Invoice No': '',
                'Invoice Date': '',
                'Model': '',
                'Status': 'INFO',
                'Mismatch Field': 'Success',
                'Master Sheet Value': '',
                'Extracted Invoice Value': '',
                'Item Report Value': '',
                'Invoice CTH': '',
                'Item CTH': '',
                'Invoice Desc': '',
                'Item Desc': '',
                'Remarks': 'Successfully verified. No mismatches found.'
            })

    logger.info("Comparison complete. %d mismatch(es) found.", len(actual_mismatches))
    
    # Calculate summary counts using the merge indicator directly.
    # 'both' = row exists in both Item Report and Invoice (matched pair).
    # 'left_only' = in Item Report but not Invoice.
    # 'right_only' = in Invoice but not Item Report.
    merge_counts = merged['_merge'].value_counts()
    both_count = merge_counts.get('both', 0)
    
    # Matched = rows that paired successfully between the two files
    matched_item_count = both_count
    matched_ext_count = both_count

    return {
        'df': pd.DataFrame(mismatches),
        'item_total': len(df_item),
        'ext_total': len(df_ext),
        'item_matched': matched_item_count,
        'ext_matched': matched_ext_count
    }


def export_to_excel(df_out: pd.DataFrame, save_path: str) -> None:
    """Write mismatch DataFrame to a formatted Excel file."""
    from openpyxl.styles import Alignment, Border, Side
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    df_exp = df_out.copy()
    
    if 'Status' in df_exp.columns:
        df_exp = df_exp[df_exp['Status'] != 'CTH_ALERT']
        statuses = df_exp['Status'].tolist()
        df_exp = df_exp.drop(columns=['Status'])
    else:
        statuses = []

    wb = Workbook(); ws = wb.active; ws.title = "Mismatch Report"
    hdr_fill = PatternFill(start_color='1F3F6E', end_color='1F3F6E', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF')
    mismatch_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    bold_val_font = Font(bold=True)

    # Get column indices (1-based) for columns that should be bolded
    bold_cols_names = {"Master Sheet Value", "Extracted Invoice Value", "Item Report Value"}
    bold_cols_idx = {i + 1 for i, col in enumerate(df_exp.columns) if col in bold_cols_names}

    ws.append(list(df_exp.columns))
    for cell in ws[1]: 
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.freeze_panes = 'A2'

    for r_idx, row_vals in enumerate(dataframe_to_rows(df_exp, index=False, header=False), 2):
        ws.append(row_vals)
        status = str(statuses[r_idx - 2]).upper().strip() if (r_idx - 2) < len(statuses) else ""
        
        for c in range(1, len(row_vals)+1):
            cell = ws.cell(row=r_idx, column=c)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            if status == "MISMATCH":
                cell.fill = mismatch_fill
            if c in bold_cols_idx:
                cell.font = bold_val_font

    for col_cells in ws.columns:
        mx = max(len(str(c.value or '')) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(mx+2, 120)

    # Add AutoFilter to all columns
    ws.auto_filter.ref = ws.dimensions

    wb.save(save_path)


# ═══════════════════════════════════════════════════════════════
# NAGARKOT-BRANDED GUI
# ═══════════════════════════════════════════════════════════════

_PRIMARY_BLUE = "#1F3F6E"; _ACCENT_RED = "#D8232A"; _DARK_TEXT = "#1E1E1E"
_MUTED_GRAY = "#6B7280"; _LIGHT_BG = "#F4F6F8"; _PANEL_WHITE = "#FFFFFF"
_BORDER_GRAY = "#E5E7EB"; _HOVER_BLUE = "#2A528F"; _HEADER_BG = "#D6E4F0"

def get_base_path() -> Path:
    return Path(sys._MEIPASS) if getattr(sys, 'frozen', False) else Path(__file__).parent

def _brand_button(parent, text: str, command, state=tk.NORMAL) -> tk.Button:
    btn = tk.Button(parent, text=text, command=command, state=state,
                    font=("Segoe UI",10,"bold"), fg="#FFF", bg=_PRIMARY_BLUE,
                    activebackground=_HOVER_BLUE, activeforeground="#FFF",
                    bd=0, padx=14, pady=6, cursor="hand2", relief=tk.FLAT)
    btn.bind("<Enter>", lambda e: btn.configure(bg=_HOVER_BLUE) if btn['state'] != 'disabled' else None)
    btn.bind("<Leave>", lambda e: btn.configure(bg=_PRIMARY_BLUE) if btn['state'] != 'disabled' else None)
    return btn


class Application:

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("SKODA Item Mismatch Checker — Nagarkot Forwarders Pvt. Ltd.")
        self.root.configure(bg=_LIGHT_BG)
        self.root.state("zoomed")
        self.root.minsize(1024, 600)

        self.master_path = ""; self.item_path = ""; self.ext_path = ""
        self.df_result: pd.DataFrame | None = None

        self._build_header()
        self._build_body()
        self._build_footer()

    # ── Header ─────────────────────────────────────────────────
    def _build_header(self) -> None:
        header = tk.Frame(self.root, bg=_PANEL_WHITE, bd=0, highlightthickness=1, highlightbackground=_BORDER_GRAY)
        header.pack(fill=tk.X, side=tk.TOP); header.pack_propagate(False); header.configure(height=85)
        try:
            self._logo_img = tk.PhotoImage(file=str(get_base_path() / "logo.png"))
            factor = max(1, self._logo_img.height() // 20)
            self._logo_img = self._logo_img.subsample(factor)
            tk.Label(header, image=self._logo_img, bg=_PANEL_WHITE).pack(side=tk.LEFT, padx=20)
        except Exception:
            tk.Label(header, text="Nagarkot", font=("Segoe UI",12,"bold"), fg=_PRIMARY_BLUE, bg=_PANEL_WHITE).pack(side=tk.LEFT, padx=20)
        tf = tk.Frame(header, bg=_PANEL_WHITE); tf.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        tk.Label(tf, text="SKODA Item Mismatch Checker", font=("Segoe UI",18,"bold"), fg=_PRIMARY_BLUE, bg=_PANEL_WHITE).pack()
        tk.Label(tf, text="Automated Invoice & Master Sheet Reconciliation", font=("Segoe UI",10), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack()

    # ── Body ───────────────────────────────────────────────────
    def _build_body(self) -> None:
        body = tk.Frame(self.root, bg=_LIGHT_BG)
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=(12,4))

        # Row 1: Toolbar
        toolbar = tk.Frame(body, bg=_LIGHT_BG); toolbar.pack(fill=tk.X, pady=(0,6))
        self.btn_master = _brand_button(toolbar, "📂 Master Sheet", self._pick_master); self.btn_master.pack(side=tk.LEFT)
        self.btn_item = _brand_button(toolbar, "📂 Item Report", self._pick_item); self.btn_item.pack(side=tk.LEFT, padx=(8,0))
        self.btn_ext = _brand_button(toolbar, "📂 Extracted Invoice", self._pick_ext); self.btn_ext.pack(side=tk.LEFT, padx=(8,0))

        self.btn_run = tk.Button(toolbar, text="▶  Run", font=("Segoe UI",10,"bold"),
                                 fg="#FFF", bg="#A0AAB5", activebackground="#b71c1c", activeforeground="#FFF",
                                 disabledforeground="#FFFFFF", bd=0, padx=18, pady=7, cursor="hand2", state=tk.DISABLED, command=self._run)
        self.btn_run.pack(side=tk.LEFT, padx=(16,0))
        self.btn_run.bind("<Enter>", lambda e: self.btn_run.configure(bg="#b71c1c") if str(self.btn_run['state']) != 'disabled' else None)
        self.btn_run.bind("<Leave>", lambda e: self.btn_run.configure(bg=_ACCENT_RED) if str(self.btn_run['state']) != 'disabled' else None)

        self.btn_export = _brand_button(toolbar, "💾  Export Excel", self._export, state=tk.DISABLED)
        self.btn_export.pack(side=tk.RIGHT)

        self.btn_clear = tk.Button(toolbar, text="🔄 Clear All", font=("Segoe UI",10),
                                  fg=_DARK_TEXT, bg="#E5E7EB", activebackground="#D1D5DB",
                                  bd=0, padx=12, pady=6, cursor="hand2", command=self._clear_all)
        self.btn_clear.pack(side=tk.RIGHT, padx=8)

        self.status_var = tk.StringVar(value="Select all three Excel files to begin")
        self.status_lbl = tk.Label(toolbar, textvariable=self.status_var, font=("Segoe UI",10,"bold"),
                                   fg=_PRIMARY_BLUE, bg=_LIGHT_BG, anchor=tk.E, padx=12)
        self.status_lbl.pack(side=tk.RIGHT, padx=(0,8))

        # Row 2: Summary bar (hidden initially)
        self._summary_frame = tk.Frame(body, bg=_LIGHT_BG); self._summary_frame.pack(fill=tk.X, pady=(0,4))
        self.summary_bar = tk.Label(self._summary_frame, text="", font=("Segoe UI",10,"bold"), padx=12, pady=6, anchor=tk.W)

        # Row 3: Table area
        self._table_frame = tk.Frame(body, bg=_PANEL_WHITE, bd=1, relief=tk.SOLID)
        self._table_frame.pack(fill=tk.BOTH, expand=True)
        self.tree = None

    # ── Footer ─────────────────────────────────────────────────
    def _build_footer(self) -> None:
        ft = tk.Frame(self.root, bg=_PANEL_WHITE, height=28, highlightthickness=1, highlightbackground=_BORDER_GRAY)
        ft.pack(fill=tk.X, side=tk.BOTTOM); ft.pack_propagate(False)
        tk.Label(ft, text="Nagarkot Forwarders Pvt. Ltd. ©", font=("Segoe UI",8), fg=_MUTED_GRAY, bg=_PANEL_WHITE).pack(side=tk.LEFT, padx=12)

    def _clear_all(self) -> None:
        """Reset all file paths, UI elements, and data."""
        self.master_path = ""
        self.item_path = ""
        self.ext_path = ""
        self.df_result = None
        self.btn_master.configure(text="📂 Master Sheet")
        self.btn_item.configure(text="📂 Item Report")
        self.btn_ext.configure(text="📂 Extracted Invoice")
        self.btn_run.configure(state=tk.DISABLED, bg="#A0AAB5", text="▶  Run")
        self.btn_export.configure(state=tk.DISABLED)
        self.status_var.set("Select all three Excel files to begin")
        self._clear_table()
        if hasattr(self, 'summary_bar'):
            self.summary_bar.pack_forget()
        logger.info("Application state reset via Clear All")

    # ── Status helpers ─────────────────────────────────────────
    def _set_status(self, text: str, color: str = _PRIMARY_BLUE) -> None:
        self.status_var.set(text); self.status_lbl.configure(fg=color); self.root.update_idletasks()

    def _check_ready(self) -> None:
        """Enable/disable Run button based on whether all 3 files are selected."""
        if self.master_path and self.item_path and self.ext_path:
            self.btn_run.configure(state=tk.NORMAL, bg=_ACCENT_RED)
            self._set_status("All files selected — click Run")
        else:
            self.btn_run.configure(state=tk.DISABLED, bg="#A0AAB5")

    # ── File pickers ───────────────────────────────────────────
    def _pick_master(self) -> None:
        p = filedialog.askopenfilename(title="Select Master Sheet", filetypes=[("Data Files","*.xlsx *.xls *.csv")])
        if p:
            self.master_path = p
            self.btn_master.configure(text=f"✅ {Path(p).name[:30]}")
            self._set_status(f"Master: {Path(p).name}")
            self._check_ready()

    def _pick_item(self) -> None:
        p = filedialog.askopenfilename(title="Select Item Report", filetypes=[("Data Files","*.xlsx *.xls *.csv")])
        if p:
            self.item_path = p
            self.btn_item.configure(text=f"✅ {Path(p).name[:30]}")
            self._set_status(f"Item Report: {Path(p).name}")
            self._check_ready()

    def _pick_ext(self) -> None:
        p = filedialog.askopenfilename(title="Select Extracted Invoice", filetypes=[("Data Files","*.xlsx *.xls *.csv")])
        if p:
            self.ext_path = p
            self.btn_ext.configure(text=f"✅ {Path(p).name[:30]}")
            self._set_status(f"Extracted Invoice: {Path(p).name}")
            self._check_ready()

    # ── Run comparison ─────────────────────────────────────────
    def _run(self) -> None:
        """Run comparison with visual processing feedback."""
        self.btn_run.configure(state=tk.DISABLED, text="⌛ Processing...", bg="#9e191e")
        self.btn_export.configure(state=tk.DISABLED)
        self.root.update()
        try:
            self._set_status("Processing...", _PRIMARY_BLUE)
            res = perform_comparison(self.master_path, self.item_path, self.ext_path)
            self.df_result = res['df']
            
            summary_msg = f"✅ Success! Matched: Item {res['item_matched']}/{res['item_total']} | Inv {res['ext_matched']}/{res['ext_total']}"
            
            if self.df_result is None or self.df_result.empty:
                self._show_summary(f"{summary_msg}. No mismatches found!", "#E8F5E9", "#2E7D32")
                self._set_status("All matched — 0 mismatches", "#2E7D32")
                self._clear_table()
            else:
                if 'Status' in self.df_result.columns:
                    mismatch_count = len(self.df_result[~self.df_result['Status'].isin(['CTH_ALERT', 'INFO'])])
                else:
                    mismatch_count = len(self.df_result)
                
                if mismatch_count == 0:
                    self._show_summary(f"{summary_msg}. No mismatches found!", "#E8F5E9", "#2E7D32")
                    self._set_status("All matched — 0 mismatches", "#2E7D32")
                    self.btn_export.configure(state=tk.DISABLED)
                else:
                    self._show_summary(f"{summary_msg} | ⚠ {mismatch_count} mismatch(es) found.", "#FFF8E1", "#F57F17")
                    self._set_status(f"{mismatch_count} mismatch(es) found", _ACCENT_RED)
                    self.btn_export.configure(state=tk.NORMAL)
                
                self._render_table()
        except Exception as e:
            logger.error("Comparison failed: %s\n%s", e, traceback.format_exc())
            messagebox.showerror("Error", f"{e}\n\nSee mismatch_checker.log for details.")
            self._set_status("Error during comparison", _ACCENT_RED)
        finally:
            self.btn_run.configure(state=tk.NORMAL, text="▶  Run", bg=_ACCENT_RED)

    # ── Summary bar ────────────────────────────────────────────
    def _show_summary(self, text: str, bg: str, fg: str) -> None:
        self.summary_bar.configure(text=text, bg=bg, fg=fg)
        self.summary_bar.pack(fill=tk.X)

    # ── Table rendering (BE Extractor style) ───────────────────
    def _clear_table(self) -> None:
        for w in self._table_frame.winfo_children(): w.destroy()
        self.tree = None

    def _render_table(self) -> None:
        self._clear_table()
        df = self.df_result
        if df is None or df.empty: return

        style = ttk.Style(); style.theme_use("clam")
        style.configure("MM.Treeview", background=_PANEL_WHITE, foreground=_DARK_TEXT,
                         fieldbackground=_PANEL_WHITE, font=("Segoe UI",10), rowheight=28)
        style.configure("MM.Treeview.Heading", background=_HEADER_BG, foreground=_PRIMARY_BLUE,
                         font=("Segoe UI",9,"bold"))
        style.map("MM.Treeview", background=[("selected",_HOVER_BLUE)], foreground=[("selected","#FFF")])

        display_cols = [c for c in df.columns if c != 'Status']
        
        # Set up grid layout on the table frame
        self._table_frame.grid_rowconfigure(0, weight=1)
        self._table_frame.grid_columnconfigure(0, weight=1)
        
        xscroll = ttk.Scrollbar(self._table_frame, orient=tk.HORIZONTAL)
        yscroll = ttk.Scrollbar(self._table_frame, orient=tk.VERTICAL)
        tree = ttk.Treeview(self._table_frame, columns=display_cols, show="headings", style="MM.Treeview",
                            xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
        xscroll.configure(command=tree.xview); yscroll.configure(command=tree.yview)
        
        # Grid placement
        tree.grid(row=0, column=0, sticky='nsew')
        yscroll.grid(row=0, column=1, sticky='ns')
        xscroll.grid(row=1, column=0, sticky='ew')
        
        self.tree = tree

        col_widths = {'Job No':90,'Job Date':90,'Invoice No':120,'Invoice Date':100,'Model':140,
                      'Mismatch Field':140,'Master Sheet Value':160,
                      'Extracted Invoice Value':180,'Item Report Value':160,
                      'Invoice CTH': 100, 'Item CTH': 100,
                      'Invoice Desc': 180, 'Item Desc': 180,
                      'Remarks':350}
        
        for c in display_cols:
            if c == 'Remarks':
                # Calculate max length of the strings in this column + the header name itself
                max_str_len = max([len(str(val)) for val in df[c].dropna()] + [len(c)])
                # Estimate width: roughly 8 pixels per character (Segoe UI 10) + 30px padding
                estimated_width = max_str_len * 8 + 30
                # Use the larger of the predefined width or the calculated width, cap at 1200px
                w = min(max(col_widths.get(c, 350), estimated_width), 1200)
            else:
                w = col_widths.get(c, 130)
            
            tree.heading(c, text=c, anchor=tk.CENTER)
            tree.column(c, width=int(w), minwidth=60, anchor=tk.CENTER, stretch=False)

        for idx, (_, row) in enumerate(df.iterrows()):
            vals = [str(row[c]) if pd.notnull(row[c]) else "" for c in display_cols]
            status = str(row.get('Status', '')).upper().strip()
            if status == "INFO": tag = "info"
            elif status == "CTH_ALERT": tag = "cth_alert"
            else: tag = "mismatch"
            tree.insert("", tk.END, values=vals, tags=(tag,))

        tree.tag_configure("mismatch", background="#FFEBEE")
        tree.tag_configure("info", background="#FFFFFF")
        tree.tag_configure("cth_alert", background="#FFFFFF", font=("Segoe UI", 10, "bold"))

        # Bind double-click for editing
        tree.bind("<Double-1>", self._on_double_click_cell)

    def _on_double_click_cell(self, event):
        """Identify cell on double-click and spawn an Entry for editing."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell": return

        column = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        if not item_id: return

        # Get column index and name
        col_index = int(column[1:]) - 1
        col_name = self.tree['columns'][col_index]

        # Get cell coordinates
        bbox = self.tree.bbox(item_id, column)
        if not bbox: return
        x, y, w, h = bbox

        # Current value
        current_val = self.tree.item(item_id, 'values')[col_index]

        # Create entry
        entry = tk.Entry(self._table_frame, font=("Segoe UI", 10))
        entry.insert(0, current_val)
        entry.select_range(0, tk.END)
        entry.focus_set()

        # Place entry
        entry.place(x=x, y=y, width=w, height=h)

        def save_edit(event=None):
            new_val = entry.get()
            # Update Treeview
            vals = list(self.tree.item(item_id, 'values'))
            vals[col_index] = new_val
            self.tree.item(item_id, values=vals)

            # Update DataFrame
            # Find index in df_result using iid or by matching other key values if iid isn't reliable
            # For simplicity, since the tree follows df_result order:
            row_idx = self.tree.index(item_id)
            if self.df_result is not None:
                self.df_result.iloc[row_idx, col_index] = new_val

            entry.destroy()

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", lambda e: entry.destroy())
        entry.bind("<Escape>", lambda e: entry.destroy())

    # ── Export ─────────────────────────────────────────────────
    def _export(self) -> None:
        if self.df_result is None or self.df_result.empty:
            messagebox.showinfo("Info", "No mismatches to export."); return
        if 'Status' in self.df_result.columns and len(self.df_result[self.df_result['Status'] != 'CTH_ALERT']) == 0:
            messagebox.showinfo("Info", "Only CTH Alerts are present. No mismatches to export.")
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = filedialog.asksaveasfilename(title="Save Mismatch Report", defaultextension=".xlsx",
                                                  initialfile=f"comparison_mismatch_report_{timestamp}.xlsx",
                                                  filetypes=[("Excel","*.xlsx")])
        if not save_path: return
        try:
            import os
            export_to_excel(self.df_result, save_path)
            self._set_status("Exported successfully", "#2E7D32")
            if messagebox.askyesno("Success", f"Report saved to:\n{save_path}\n\nDo you want to open it now?"):
                os.startfile(save_path)
        except Exception as e:
            messagebox.showerror("Export Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    Application(root)
    root.mainloop()
